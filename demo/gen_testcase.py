from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "供应商搭赠测试用例"

header_fill = PatternFill("solid", start_color="2F5496", end_color="2F5496")
module1_fill = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
module2_fill = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")

thin = Side(style='thin', color="AAAAAA")
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

title_font = Font(name="Arial", bold=True, size=14, color="2F5496")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
normal_font = Font(name="Arial", size=10)
bold_font = Font(name="Arial", bold=True, size=10)

ws.merge_cells("A1:H1")
ws["A1"] = "供应商搭赠 & 申请单审核搭赠 测试用例"
ws["A1"].font = title_font
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:H2")
ws["A2"] = "功能模块：门户端-供应商搭赠 | 门户端-申请单审核搭赠          编写日期：2026-04-08"
ws["A2"].font = Font(name="Arial", size=9, color="666666", italic=True)
ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 18

ws.append([""] * 8)
ws.row_dimensions[3].height = 5

headers = ["用例编号", "所属模块", "用例标题", "前置条件", "测试步骤", "预期结果", "优先级", "用例类型"]
ws.append(headers)
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col)
    cell.value = h
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[4].height = 22

test_cases = [
    (
        "TC-001", "门户端-供应商搭赠", "供应商搭赠列表展示",
        "已登录门户端，供应商账号存在搭赠配置",
        "1. 进入供应商搭赠页面",
        "正确展示搭赠列表，含物料编码、物料名称、规格、供应商、生产商、采购单位等列，数据显示完整",
        "P1", "功能测试"
    ),
    (
        "TC-002", "门户端-供应商搭赠", '赠品"是否赠品"列独立显示',
        "申请单中含赠品物料",
        "1. 创建采购申请单\n2. 编辑采购申请单\n3. 查看采购申请单\n4. 分别检查物料列表",
        '三个场景下，"是否赠品"列均不与其他列合并，单独展示，赠品行标识清晰可辨',
        "P1", "功能测试"
    ),
    (
        "TC-003", "门户端-供应商搭赠", "可用赠品数量计算正确",
        "物料有搭赠配置，历史已使用部分赠品",
        "1. 打开含赠品的采购申请单\n2. 查看赠品物料行的可用赠品数量",
        "可用赠品数量 = 累计可用量 - 当前申请单已使用量，数值计算正确",
        "P1", "功能测试"
    ),
    (
        "TC-004", "门户端-供应商搭赠", '是否赠品切换时数量联动更新',
        "采购申请单处于编辑状态，含赠品物料",
        '1. 将某物料"是否赠品"由"是"改为"否"\n2. 观察可用赠品数量变化\n3. 再改回"是"，再次观察',
        '"是否赠品"切换后，可用赠品数量实时联动更新；改回"是"后数量恢复正确值',
        "P1", "功能测试"
    ),
    (
        "TC-005", "门户端-申请单审核搭赠", "审核页面赠品列正常展示",
        "审核员账号登录，有待审核的含赠品申请单",
        "1. 打开审核页面\n2. 查看物料列表",
        '物料列表中展示"可用赠品数量"列，赠品行有"使用赠品"操作入口',
        "P1", "功能测试"
    ),
    (
        "TC-006", "门户端-申请单审核搭赠", "单条赠品使用弹窗-正常填写并确认",
        "审核页面，物料101013可用赠品数量为100，下单数量200",
        '1. 点击赠品行的"使用赠品"按钮\n2. 确认弹窗展示：物料编码、名称、规格、下单数量200、可用赠品数量100\n3. 输入使用赠品数量（如50）\n4. 点击"确定"',
        "弹窗关闭；该行可用赠品数量更新为50；页面有操作成功提示",
        "P1", "功能测试"
    ),
    (
        "TC-007", "门户端-申请单审核搭赠", "单条赠品使用弹窗-输入超出可用数量",
        "可用赠品数量为100",
        '1. 打开"使用赠品"弹窗\n2. 输入使用数量150\n3. 点击"确定"',
        '提示错误："使用赠品数量不能超过可用赠品数量100"，不允许提交，弹窗保持打开',
        "P1", "边界测试"
    ),
    (
        "TC-008", "门户端-申请单审核搭赠", "单条赠品使用弹窗-输入0或负数",
        "弹窗已打开",
        '1. 输入0，点击"确定"\n2. 输入-1，点击"确定"',
        "两种情况均提示校验错误：使用数量须大于0，不允许提交",
        "P1", "边界测试"
    ),
    (
        "TC-009", "门户端-申请单审核搭赠", "单条赠品使用弹窗-使用数量为空",
        "弹窗已打开，使用赠品数量输入框为空",
        '1. 不填写使用赠品数量\n2. 点击"确定"',
        "提示必填校验错误，不允许提交，弹窗保持打开",
        "P2", "边界测试"
    ),
    (
        "TC-010", "门户端-申请单审核搭赠", "单条赠品使用弹窗-取消操作",
        "弹窗已打开并已填写使用数量",
        '1. 填写使用数量（如50）\n2. 点击"取消"\n3. 再次以点击"×"关闭验证',
        "弹窗关闭，可用赠品数量保持不变，数据未被修改",
        "P2", "功能测试"
    ),
    (
        "TC-011", "门户端-申请单审核搭赠", "批量使用赠品-确认全部使用",
        "审核页面存在多条赠品物料，各有不同可用赠品数量",
        '1. 点击"批量使用赠品"按钮\n2. 确认弹窗提示内容\n3. 点击"确定"',
        "所有赠品行使用数量自动填充为各自的可用赠品数量；全部使用成功；操作成功提示",
        "P1", "功能测试"
    ),
    (
        "TC-012", "门户端-申请单审核搭赠", "批量使用赠品-取消操作",
        "批量使用赠品确认弹窗已打开",
        '1. 点击"取消"',
        "弹窗关闭，所有赠品使用数量保持不变",
        "P2", "功能测试"
    ),
    (
        "TC-013", "门户端-申请单审核搭赠", "批量使用赠品-所有赠品可用数量为0",
        "申请单所有赠品物料的可用赠品数量均为0",
        "1. 查看批量使用赠品按钮状态\n2. 尝试点击（若可点击）",
        '按钮置灰不可点击，或点击后提示"当前无可用赠品"',
        "P2", "边界测试"
    ),
    (
        "TC-014", "门户端-申请单审核搭赠", "审核通过后赠品数量同步更新",
        "审核员已对赠品使用数量进行设置",
        "1. 审核员完成赠品使用设置\n2. 提交审核并通过\n3. 查看供应商搭赠统计页面",
        "累计可用量正确扣减已使用赠品数量；搭赠统计数据同步更新，数据一致",
        "P1", "集成测试"
    ),
]

row_num = 5
for tc in test_cases:
    ws.append(tc)
    ws.row_dimensions[row_num].height = 60

    fill = module1_fill if tc[1] == "门户端-供应商搭赠" else module2_fill

    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.font = bold_font if col in (1, 2, 3) else normal_font
        cell.alignment = Alignment(
            horizontal="center" if col in (1, 2, 7, 8) else "left",
            vertical="center", wrap_text=True
        )
        cell.border = thin_border
        cell.fill = fill

    p_cell = ws.cell(row=row_num, column=7)
    if tc[6] == "P1":
        p_cell.font = Font(name="Arial", bold=True, size=10, color="C00000")
    else:
        p_cell.font = Font(name="Arial", bold=True, size=10, color="FF6600")

    row_num += 1

col_widths = [10, 22, 30, 32, 40, 44, 8, 12]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A5"

out_path = "/Users/freya/Desktop/code/demo/供应商搭赠测试用例.xlsx"
wb.save(out_path)
print("saved:", out_path)
