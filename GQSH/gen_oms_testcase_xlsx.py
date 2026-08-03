# -*- coding: utf-8 -*-
"""生成 OMS 各子模块 xlsx 测试用例文件"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'OmstestCase')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 通用样式
HEADER_FONT = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
CELL_ALIGN = Alignment(vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

HEADERS = ['用例编号', '所属模块', '用例名称', '接口地址', '请求方式', '请求参数', '预期结果', '优先级']
COL_WIDTHS = [12, 16, 28, 42, 10, 50, 35, 8]


def create_workbook(module_name, test_cases):
    wb = Workbook()
    ws = wb.active
    ws.title = module_name

    # 写入表头
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # 写入数据
    for row_idx, case in enumerate(test_cases, 2):
        for col_idx, value in enumerate(case, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER

    # 设置列宽
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    return wb


# ── 各模块测试用例数据 ──

modules = {
    '数据看板': [
        ['TC-Dashboard-001', '数据看板', '查询概览统计数据', '/api/oms-admin/api/dashboard/overview', 'POST',
         'startTime, endTime', '返回 success=True，包含概览统计数据', 'P0'],
        ['TC-Dashboard-002', '数据看板', '查询销售趋势数据', '/api/oms-admin/api/dashboard/salesTrend', 'POST',
         'startTime, endTime, granularity', '返回 success=True，包含趋势图表数据', 'P1'],
        ['TC-Dashboard-003', '数据看板', '查询订单统计数据', '/api/oms-admin/api/dashboard/orderStatistics', 'POST',
         'startTime, endTime', '返回 success=True，包含订单统计信息', 'P1'],
    ],
    '三方数据': [
        ['TC-ThirdParty-001', '三方数据', '查询三方数据同步列表', '/api/oms-admin/api/thirdPartyData/syncList', 'POST',
         'startTime, endTime, sourceType, syncStatus, page, limit', '返回分页列表，totalCount>=0', 'P0'],
        ['TC-ThirdParty-002', '三方数据', '查询三方数据同步详情', '/api/oms-admin/api/thirdPartyData/syncDetail', 'POST',
         'syncId/id（来自同步列表）, startTime, endTime', '返回 success=True，包含同步详情', 'P1'],
        ['TC-ThirdParty-003', '三方数据', '查询渠道数据列表', '/api/oms-admin/api/thirdPartyData/channelList', 'POST',
         'channelCode, channelName, page, limit', '返回渠道分页列表', 'P1'],
    ],
    '调拨管理': [
        ['TC-Transfer-001', '调拨管理', '查询调拨单列表', '/api/oms-admin/api/transferOrder/page', 'POST',
         'startTime, endTime, transferOrderNo, fromWarehouseCode, toWarehouseCode, status, page, limit',
         '返回调拨单分页列表', 'P0'],
        ['TC-Transfer-002', '调拨管理', '查询调拨单详情', '/api/oms-admin/api/transferOrder/detail', 'POST',
         'transferOrderNo（来自调拨单列表）, startTime, endTime', '返回 success=True，包含调拨单详情', 'P1'],
        ['TC-Transfer-003', '调拨管理', '查询调拨入库单列表', '/api/oms-admin/api/transferInbound/page', 'POST',
         'startTime, endTime, inboundOrderNo, transferOrderNo, status, page, limit',
         '返回调拨入库单分页列表', 'P0'],
        ['TC-Transfer-004', '调拨管理', '查询调拨出库单列表', '/api/oms-admin/api/transferOutbound/page', 'POST',
         'startTime, endTime, outboundOrderNo, transferOrderNo, status, page, limit',
         '返回调拨出库单分页列表', 'P0'],
    ],
    '区域实时库存': [
        ['TC-RegionalInv-001', '区域实时库存', '查询区域库存列表', '/api/oms-admin/api/regionalInventory/page', 'POST',
         'regionCode, regionName, materialCode, materialName, warehouseCode, page, limit',
         '返回区域库存分页列表', 'P0'],
        ['TC-RegionalInv-002', '区域实时库存', '查询区域库存汇总', '/api/oms-admin/api/regionalInventory/summary', 'POST',
         'regionCode, startTime, endTime', '返回 success=True，包含汇总数据', 'P1'],
        ['TC-RegionalInv-003', '区域实时库存', '查询库存明细', '/api/oms-admin/api/regionalInventory/detail', 'POST',
         'regionCode, warehouseCode, materialCode, page, limit', '返回库存明细分页列表', 'P1'],
    ],
    'F2B': [
        ['TC-F2B-001', 'F2B', '查询F2B订单列表', '/api/oms-admin/api/f2b/order/page', 'POST',
         'startTime, endTime, f2bOrderNo, thirdOrderNo, status, page, limit',
         '返回F2B订单分页列表', 'P0'],
        ['TC-F2B-002', 'F2B', '查询F2B订单详情', '/api/oms-admin/api/f2b/order/detail', 'POST',
         'f2bOrderNo（来自订单列表）, startTime, endTime', '返回 success=True，包含订单详情', 'P1'],
        ['TC-F2B-003', 'F2B', '查询F2B发货列表', '/api/oms-admin/api/f2b/delivery/page', 'POST',
         'startTime, endTime, f2bOrderNo, deliveryNo, status, page, limit',
         '返回F2B发货分页列表', 'P0'],
    ],
    '仓储': [
        ['TC-Warehouse-001', '仓储', '查询库存调拨列表', '/api/oms-admin/api/warehouse/stockTransfer/page', 'POST',
         'startTime, endTime, transferNo, fromWarehouse, toWarehouse, status, page, limit',
         '返回库存调拨分页列表', 'P0'],
        ['TC-Warehouse-002', '仓储', '查询仓库库存列表', '/api/oms-admin/api/warehouse/inventory/page', 'POST',
         'warehouseCode, materialCode, materialName, page, limit',
         '返回仓库库存分页列表', 'P0'],
        ['TC-Warehouse-003', '仓储', '查询入库单列表', '/api/oms-admin/api/warehouse/stockIn/page', 'POST',
         'startTime, endTime, stockInNo, warehouseCode, status, page, limit',
         '返回入库单分页列表', 'P0'],
        ['TC-Warehouse-004', '仓储', '查询出库单列表', '/api/oms-admin/api/warehouse/stockOut/page', 'POST',
         'startTime, endTime, stockOutNo, warehouseCode, status, page, limit',
         '返回出库单分页列表', 'P0'],
        ['TC-Warehouse-005', '仓储', '查询仓库列表', '/api/oms-admin/api/warehouse/list', 'POST',
         'warehouseCode, warehouseName, page, limit', '返回仓库分页列表', 'P1'],
    ],
    '对码表': [
        ['TC-CodeMap-001', '对码表', '查询SKU对码列表', '/api/oms-admin/api/codeMapping/skuMapping/page', 'POST',
         'internalSkuCode, internalSkuName, externalSkuCode, channelCode, page, limit',
         '返回SKU对码分页列表', 'P0'],
        ['TC-CodeMap-002', '对码表', '查询仓库对码列表', '/api/oms-admin/api/codeMapping/warehouseMapping/page', 'POST',
         'internalWarehouseCode, internalWarehouseName, externalWarehouseCode, channelCode, page, limit',
         '返回仓库对码分页列表', 'P0'],
        ['TC-CodeMap-003', '对码表', '查询渠道列表', '/api/oms-admin/api/codeMapping/channel/page', 'POST',
         'channelCode, channelName, page, limit', '返回渠道分页列表', 'P1'],
    ],
    '基础数据': [
        ['TC-BasicData-001', '基础数据', '查询供应商列表', '/api/oms-admin/api/basicData/supplier/page', 'POST',
         'supplierCode, supplierName, status, page, limit', '返回供应商分页列表', 'P0'],
        ['TC-BasicData-002', '基础数据', '查询商品物料列表', '/api/oms-admin/api/basicData/material/page', 'POST',
         'materialCode, materialName, categoryCode, page, limit', '返回商品物料分页列表', 'P0'],
        ['TC-BasicData-003', '基础数据', '查询仓库列表', '/api/oms-admin/api/basicData/warehouse/page', 'POST',
         'warehouseCode, warehouseName, warehouseType, page, limit', '返回仓库分页列表', 'P0'],
        ['TC-BasicData-004', '基础数据', '查询商品分类列表', '/api/oms-admin/api/basicData/category/page', 'POST',
         'categoryCode, categoryName, parentCode, page, limit', '返回商品分类分页列表', 'P1'],
    ],
    '策略配置': [
        ['TC-Strategy-001', '策略配置', '查询订单路由策略列表', '/api/oms-admin/api/strategy/routing/page', 'POST',
         'strategyName, strategyType, status, page, limit', '返回路由策略分页列表', 'P0'],
        ['TC-Strategy-002', '策略配置', '查询分仓策略列表', '/api/oms-admin/api/strategy/allocation/page', 'POST',
         'strategyName, warehouseCode, regionCode, status, page, limit', '返回分仓策略分页列表', 'P0'],
        ['TC-Strategy-003', '策略配置', '查询库存策略列表', '/api/oms-admin/api/strategy/inventory/page', 'POST',
         'strategyName, warehouseCode, materialCode, status, page, limit', '返回库存策略分页列表', 'P1'],
    ],
    '辅助功能': [
        ['TC-Auxiliary-001', '辅助功能', '查询操作日志列表', '/api/oms-admin/api/auxiliary/operationLog/page', 'POST',
         'startTime, endTime, operator, module, action, page, limit', '返回操作日志分页列表', 'P0'],
        ['TC-Auxiliary-002', '辅助功能', '查询导出任务列表', '/api/oms-admin/api/auxiliary/exportTask/page', 'POST',
         'startTime, endTime, taskName, status, page, limit', '返回导出任务分页列表', 'P1'],
        ['TC-Auxiliary-003', '辅助功能', '查询系统消息列表', '/api/oms-admin/api/auxiliary/message/page', 'POST',
         'startTime, endTime, messageType, isRead, page, limit', '返回系统消息分页列表', 'P1'],
    ],
    '设置': [
        ['TC-Settings-001', '设置', '查询用户列表', '/api/oms-admin/api/settings/user/page', 'POST',
         'username, realName, status, page, limit', '返回用户分页列表', 'P0'],
        ['TC-Settings-002', '设置', '查询角色列表', '/api/oms-admin/api/settings/role/page', 'POST',
         'roleName, roleCode, status, page, limit', '返回角色分页列表', 'P0'],
        ['TC-Settings-003', '设置', '查询数据字典列表', '/api/oms-admin/api/settings/dict/page', 'POST',
         'dictType, dictName, page, limit', '返回数据字典分页列表', 'P1'],
        ['TC-Settings-004', '设置', '查询系统配置', '/api/oms-admin/api/settings/systemConfig', 'POST',
         '{}', '返回 success=True，包含系统配置信息', 'P1'],
    ],
    '数据导入': [
        ['TC-DataImport-001', '数据导入', '查询导入任务列表', '/api/oms-admin/api/dataImport/task/page', 'POST',
         'startTime, endTime, taskName, importType, status, page, limit', '返回导入任务分页列表', 'P0'],
        ['TC-DataImport-002', '数据导入', '查询导入模板列表', '/api/oms-admin/api/dataImport/template/page', 'POST',
         'templateName, templateType, page, limit', '返回导入模板分页列表', 'P1'],
        ['TC-DataImport-003', '数据导入', '查询导入记录列表', '/api/oms-admin/api/dataImport/record/page', 'POST',
         'startTime, endTime, importType, operator, status, page, limit', '返回导入记录分页列表', 'P0'],
    ],
}


def main():
    for module_name, cases in modules.items():
        wb = create_workbook(module_name, cases)
        file_path = os.path.join(OUTPUT_DIR, f'OMS-{module_name}-测试用例.xlsx')
        wb.save(file_path)
        print(f'已生成: {file_path}')

    print(f'\n共生成 {len(modules)} 个测试用例文件，输出目录: {OUTPUT_DIR}')


if __name__ == '__main__':
    main()
